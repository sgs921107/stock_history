#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import json
from openpyxl.workbook import Workbook
from openpyxl import load_workbook


class DataFrame2Excel(object):

    def __init__(self, xls_filename='new_file.xlsx', mode="w"):
        """
        初始化
        :param xls_filename:用于保存数据的表格文件名
        :param mode:保存方式 a:追加/w:重写
        """
        # 如果表格文件名存在，且为追加模式
        if os.path.exists(xls_filename) and mode=="a":
            # 加载表格文件
            self.book = load_workbook(xls_filename)
        else:
            # 新建表格
            self.book = Workbook()
        # 构建列标签的map，以通过数字找到列标签
        num_list = [i for i in range(1, 15)]
        letter_list = list("ABCDEFGHIJKLMN")
        self.row_map = {}
        for num, letter in zip(num_list, letter_list):
            self.row_map[num] = letter
        # 用于保存表格数据的文件名
        self.xls_filename = xls_filename

    def _data_frame2sheet(self, df, title):
        """
        将一条json数据转化为一张表格
        :param df: pandas.DataFrame数据
        :return: None
        """
	# 获取列索引
        columns = df.columns.to_list()
        # 获取最大列数
        # 新建sheet
	sheet = self.book.create_sheet(title=title, index=0)
        # sheet.append(title)
        # 根据最大列数合并title行
        # sheet.merge_cells('A1:%s1' % self.row_map[len(columns)])
	sheet.append(columns)
        for index in df.index.to_list():
            sheet.append(df.loc[index].to_list())

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.save()
        if exc_type:
            print(exc_type)
            print(exc_val)
            print(exc_tb)

    def data_frame2sheet(self, df, title):
        self._data_frame2sheet(df, title)

    def data_frames2sheets(self, dfs):
        """
        将df数据数据存入多张表
        :param filenames:
        :return:
        """
        for title, df in dfs.items():
            self._data_frame2sheet(df, title)

    def save(self):
        self.book.save(self.xls_filename)


if __name__ == '__main__':
    import json
    import pandas as pd
    datafile_path = "./products_cp.json"
    with open(datafile_path) as f:
	data = (json.loads(line) for line in f)
	df = pd.DataFrame(data)
    date_group = df.groupby("date")
    date_data = date_group.get_group("2018-04-25")
    rank_data = date_data.sort_values('turnover', ascending=False)[0:50][["stock_code", "stock_abbreviation", "trading_volume", "turnover"]]
    rank_data.insert(0, "date_rank", range(1, 51))
    with DataFrame2Excel("zhongtou.xlsx", "w") as b:
	b.data_frame2sheet(rank_data, "2018-04-25")
	

"""
__title__ = ''
__author__ = 'v_sungensheng'
__mtime__ = '2018/10/23'
# code is far away from bugs with the god animal protecting
    I love animals. They taste delicious.
              ┏┓      ┏┓
            ┏┛┻━━━┛┻┓
            ┃      ☃      ┃
            ┃  ┳┛  ┗┳  ┃
            ┃      ┻      ┃
            ┗━┓      ┏━┛
                ┃      ┗━━━┓
                ┃  神兽保佑    ┣┓
                ┃　永无BUG！   ┏┛
                ┗┓┓┏━┳┓┏┛
                  ┃┫┫  ┃┫┫
                  ┗┻┛  ┗┻┛
"""
